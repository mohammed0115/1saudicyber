#!/usr/bin/env python3
"""
Generate a realistic (FICTIONAL) Saudi tech-company test dataset for 1SaudiCyber.

All data is invented — no real company, person, or secret. Interlinked across
registers/reports/policies with deliberate, logical gaps for testing evidence
extraction, classification, gap discovery, AI advisory, and RFI flows.

Output: ./Enterprise_Test_Dataset/{Policies,Registers,Reports,Evidence,Screenshots,
        Network,Risk,Assets,Incidents,Training,Third_Party,Compliance,Infrastructure}
"""
import os, datetime
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from PIL import Image, ImageDraw, ImageFont

ROOT = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), 'Enterprise_Test_Dataset')
DIRS = ['Policies','Registers','Reports','Evidence','Screenshots','Network','Risk',
        'Assets','Incidents','Training','Third_Party','Compliance','Infrastructure']
for d in DIRS:
    os.makedirs(os.path.join(ROOT, d), exist_ok=True)

# ---------------------------------------------------------------- Company facts
CO = dict(
    name_en='Nukhba Digital Technologies Co.', name_ar='شركة نُخبة التقنية الرقمية',
    short='NDT', cr='1010874321', vat='3001234567800003',
    city='Riyadh', hq='Olaya District, Riyadh', branches=['Jeddah', 'Dammam'],
    employees=320, sector='Cloud & Managed IT / FinTech SaaS',
    domain='ndt.com.sa', cloud='Microsoft Azure (UAE North / me-central-1)',
    idp='Microsoft Entra ID', dc_primary='Riyadh DC-1', dc_dr='Jeddah DR-2',
    ciso='Faisal Al-Harbi', itd='Noura Al-Qahtani', socm='Yousef Al-Dossary',
    cco='Layla Al-Mutairi', ceo='Abdulaziz Al-Rasheed',
)
TODAY = datetime.date(2026, 7, 19)
def d(days): return (TODAY - datetime.timedelta(days=days)).isoformat()

# Source-of-truth servers (referenced everywhere)
SERVERS = [
    # code, name, os, role, ip, dc, owner, critical
    ('APP-SRV-01','App Server 01','Windows Server 2022','FinTech App',      '10.10.20.11','Riyadh DC-1','App Team','High'),
    ('APP-SRV-02','App Server 02','Windows Server 2019','FinTech App (sec)','10.10.20.12','Riyadh DC-1','App Team','High'),
    ('DB-SRV-01', 'DB Server 01', 'Windows Server 2022','SQL Database',     '10.10.20.21','Riyadh DC-1','DBA Team','High'),
    ('WEB-SRV-01','Web Server 01','Ubuntu 22.04 LTS',   'Public Web/API',   '10.10.10.31','DMZ',        'Web Team','High'),
    ('AD-SRV-01', 'AD/DC 01',     'Windows Server 2022','Active Directory',  '10.10.30.5', 'Riyadh DC-1','Infra Team','High'),
    ('BKP-SRV-01','Backup Server','Ubuntu 22.04 LTS',   'Veeam Backup',     '10.10.40.9', 'Riyadh DC-1','Infra Team','High'),
    ('SIEM-SRV-01','SIEM Server', 'Ubuntu 22.04 LTS',   'Wazuh SIEM',       '10.10.40.15','Riyadh DC-1','SOC Team', 'High'),
    ('FILE-SRV-01','File Server', 'Windows Server 2019','File Shares',      '10.10.20.41','Riyadh DC-1','Infra Team','Medium'),
    ('LEGACY-SRV-09','Legacy Billing','Windows Server 2012 R2','Legacy Billing','10.10.20.99','Riyadh DC-1','(none)','Medium'),  # GAP: no backup, no owner
]

# ------------------------------------------------------------------ PDF helpers
styles = getSampleStyleSheet()
H = ParagraphStyle('H', parent=styles['Heading1'], textColor=colors.HexColor('#1f4135'), fontSize=17, spaceAfter=6)
H2 = ParagraphStyle('H2', parent=styles['Heading2'], textColor=colors.HexColor('#2a5646'), fontSize=12, spaceBefore=10, spaceAfter=3)
BODY = ParagraphStyle('B', parent=styles['BodyText'], fontSize=9.5, leading=14)
SMALL = ParagraphStyle('S', parent=styles['BodyText'], fontSize=8, textColor=colors.HexColor('#5b5c53'))
GREEN = colors.HexColor('#2a5646'); SAND = colors.HexColor('#b49a6a'); LINE = colors.HexColor('#d3d6ca')

def _header_footer(canvas, doc):
    canvas.saveState()
    canvas.setFillColor(GREEN); canvas.rect(0, A4[1]-1.15*cm, A4[0], 1.15*cm, fill=1, stroke=0)
    canvas.setFillColor(colors.white); canvas.setFont('Helvetica-Bold', 11)
    canvas.drawString(1.5*cm, A4[1]-0.78*cm, CO['short'] + ' — ' + CO['name_en'])
    canvas.setFont('Helvetica', 7.5); canvas.drawRightString(A4[0]-1.5*cm, A4[1]-0.78*cm, 'CONFIDENTIAL — Internal Use')
    canvas.setFillColor(colors.HexColor('#9aa4b2')); canvas.setFont('Helvetica', 7)
    canvas.drawString(1.5*cm, 0.8*cm, 'Fictional test data — %s' % CO['name_en'])
    canvas.drawRightString(A4[0]-1.5*cm, 0.8*cm, 'Page %d' % doc.page)
    canvas.setStrokeColor(SAND); canvas.setLineWidth(1.4); canvas.line(1.5*cm,0.62*cm,A4[0]-1.5*cm,0.62*cm)
    canvas.restoreState()

def doc_control(doc_id, version, owner, approver, classification, review_days_ago, next_review_days):
    rows = [['Document ID', doc_id, 'Version', version],
            ['Owner', owner, 'Approved by', approver],
            ['Classification', classification, 'Status', 'Approved'],
            ['Last Review', d(review_days_ago), 'Next Review', d(-next_review_days)]]
    t = Table(rows, colWidths=[3*cm, 5.6*cm, 3*cm, 4.4*cm])
    t.setStyle(TableStyle([('GRID',(0,0),(-1,-1),0.5,LINE),('FONTSIZE',(0,0),(-1,-1),8),
        ('BACKGROUND',(0,0),(0,-1),colors.HexColor('#eef2f0')),('BACKGROUND',(2,0),(2,-1),colors.HexColor('#eef2f0')),
        ('TEXTCOLOR',(0,0),(0,-1),GREEN),('TEXTCOLOR',(2,0),(2,-1),GREEN),
        ('FONTNAME',(0,0),(0,-1),'Helvetica-Bold'),('FONTNAME',(2,0),(2,-1),'Helvetica-Bold'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4)]))
    return t

def build_pdf(path, title, doc_id, version, sections, owner=None, approver=None,
              classification='Confidential', review_days_ago=95, next_review_days=270, frameworks='NCA ECC, ISO 27001'):
    owner = owner or CO['ciso']; approver = approver or CO['ceo']
    doc = SimpleDocTemplate(path, pagesize=A4, topMargin=1.6*cm, bottomMargin=1.4*cm, leftMargin=1.5*cm, rightMargin=1.5*cm)
    el = [Spacer(1,0.3*cm), Paragraph(title, H),
          Paragraph('%s · CR %s · %s' % (CO['name_en'], CO['cr'], CO['city']), SMALL),
          Spacer(1,0.25*cm), doc_control(doc_id, version, owner, approver, classification, review_days_ago, next_review_days),
          Spacer(1,0.15*cm), Paragraph('Framework alignment: %s' % frameworks, SMALL),
          HRFlowable(width='100%', color=LINE, spaceBefore=8, spaceAfter=6)]
    for htext, body in sections:
        el.append(Paragraph(htext, H2))
        for para in body:
            el.append(Paragraph(para, BODY)); el.append(Spacer(1,2))
    el += [Spacer(1,0.5*cm), HRFlowable(width='100%', color=LINE, spaceAfter=6),
           Paragraph('<b>Approval</b>', H2),
           Paragraph('Prepared by: %s (CISO) &nbsp;&nbsp; Reviewed by: %s (Compliance) &nbsp;&nbsp; Approved by: %s (CEO)'
                     % (owner, CO['cco'], approver), BODY),
           Paragraph('Signatures on file. This document is fictional test data.', SMALL)]
    doc.build(el, onFirstPage=_header_footer, onLaterPages=_header_footer)

# --------------------------------------------------------------- XLSX helper
HDR_FILL = PatternFill('solid', fgColor='2A5646'); HDR_FONT = Font(color='FFFFFF', bold=True, size=10)
THIN = Border(*[Side(style='thin', color='D9DDD3')]*4)
def build_xlsx(path, sheet, headers, rows, note=''):
    wb = Workbook(); ws = wb.active; ws.title = sheet[:31]
    r0 = 1
    if note:
        ws.cell(1,1,note).font = Font(italic=True, color='8A6A1F', size=9); r0 = 3
    for j,h in enumerate(headers,1):
        c = ws.cell(r0,j,h); c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = Alignment(horizontal='center', wrap_text=True); c.border = THIN
    for i,row in enumerate(rows, r0+1):
        for j,val in enumerate(row,1):
            c = ws.cell(i,j,val); c.border = THIN; c.font = Font(size=9); c.alignment = Alignment(vertical='center')
    for j,h in enumerate(headers,1):
        w = max(len(str(h)), *(len(str(r[j-1])) for r in rows)) if rows else len(str(h))
        ws.column_dimensions[chr(64+j) if j<=26 else 'A'+chr(64+j-26)].width = min(max(w+2, 11), 40)
    ws.freeze_panes = ws.cell(r0+1,1)
    wb.save(path)

# --------------------------------------------------------------- PNG dashboard
def png_dashboard(path, product, title, metrics, accent='#0f6cbd', status_ok=True):
    W,Hh = 1180, 720
    img = Image.new('RGB',(W,Hh),'#f3f4f6'); dr = ImageDraw.Draw(img)
    def font(sz, bold=False):
        for p in ['/usr/share/fonts/truetype/dejavu/DejaVuSans%s.ttf' % ('-Bold' if bold else ''),
                  '/usr/share/fonts/truetype/liberation/LiberationSans%s.ttf' % ('-Bold' if bold else '')]:
            if os.path.exists(p):
                try: return ImageFont.truetype(p, sz)
                except Exception: pass
        return ImageFont.load_default()
    ac = tuple(int(accent[i:i+2],16) for i in (1,3,5))
    dr.rectangle([0,0,W,64], fill=ac)
    dr.text((24,18), product, font=font(22,True), fill='white')
    dr.text((W-260,22), TODAY.isoformat(), font=font(15), fill='white')
    dr.rectangle([0,64,250,Hh], fill='#20242b')
    for i,item in enumerate(['Overview','Dashboard','Reports','Alerts','Settings','Users','Policies']):
        dr.text((22,110+i*46), item, font=font(15, i==1), fill='#e6e8ec' if i==1 else '#9aa4b2')
    dr.text((286,92), title, font=font(24,True), fill='#1f2430')
    badge = ('#1a7f37','Healthy') if status_ok else ('#b7602a','Needs Attention')
    dr.rounded_rectangle([286,132,286+150,132+30], 8, fill=badge[0]); dr.text((300,138), badge[1], font=font(14,True), fill='white')
    x,y = 286,196
    for i,(k,v,ok) in enumerate(metrics):
        cx = x + (i%3)*290; cy = y + (i//3)*150
        dr.rounded_rectangle([cx,cy,cx+265,cy+128], 12, fill='white', outline='#e2e4e8', width=1)
        dr.text((cx+18,cy+16), k, font=font(13), fill='#6b7280')
        col = '#1a7f37' if ok else '#b3261e'
        dr.text((cx+18,cy+46), str(v), font=font(34,True), fill=col)
        dr.rounded_rectangle([cx+18,cy+100,cx+18+70,cy+100+16], 6, fill='#e7f1ea' if ok else '#f7e6e3')
        dr.text((cx+26,cy+101), 'OK' if ok else 'REVIEW', font=font(11,True), fill=col)
    dr.text((286,Hh-40), 'Fictional evidence screenshot — %s — %s' % (CO['short'], product), font=font(12), fill='#9aa4b2')
    img.save(path)

# --------------------------------------------------------------- Network diagram
def png_network(path):
    W,Hh = 1500, 980
    img = Image.new('RGB',(W,Hh),'#f6f7f5'); dr = ImageDraw.Draw(img)
    def font(sz,b=False):
        for p in ['/usr/share/fonts/truetype/dejavu/DejaVuSans%s.ttf'%('-Bold' if b else ''),
                  '/usr/share/fonts/truetype/liberation/LiberationSans%s.ttf'%('-Bold' if b else '')]:
            if os.path.exists(p):
                try: return ImageFont.truetype(p,sz)
                except Exception: pass
        return ImageFont.load_default()
    dr.rectangle([0,0,W,70], fill='#2a5646'); dr.text((26,20), '%s — Network Architecture (High Level)'%CO['short'], font=font(24,True), fill='white')
    def box(x,y,w,h,label,sub='',fill='#ffffff',fg='#1f2430',outline='#356a54'):
        dr.rounded_rectangle([x,y,x+w,y+h],12,fill=fill,outline=outline,width=2)
        dr.text((x+14,y+12), label, font=font(16,True), fill=fg)
        if sub: dr.text((x+14,y+38), sub, font=font(12), fill='#5b5c53')
        return (x+w//2, y+h//2)
    def link(a,b,color='#8a8c82'):
        dr.line([a,b], fill=color, width=3)
    inet = box(650,90,200,60,'Internet','', '#eaf1ed');
    fw   = box(640,190,220,70,'Perimeter Firewall','FortiGate FW-01', '#fff5e6', outline='#b7602a')
    dmz  = box(200,320,360,120,'DMZ','WEB-SRV-01 (10.10.10.31)\nReverse Proxy / WAF', '#f6faf8')
    core = box(650,320,360,90,'Core Switches','SW-CORE-01 / SW-CORE-02', '#f6faf8')
    vpn  = box(1120,320,300,90,'VPN Gateway','Remote Access (MFA)', '#f6faf8', outline='#356a54')
    srvz = box(560,500,560,150,'Internal Server Zone','APP-SRV-01/02 · DB-SRV-01 · AD-SRV-01\nFILE-SRV-01 · LEGACY-SRV-09', '#eef4f1')
    idsvc= box(200,520,300,110,'Identity Services','Microsoft Entra ID\nConditional Access + MFA', '#eef4f1', outline='#0f6cbd')
    soc  = box(1180,520,240,110,'SOC / SIEM','SIEM-SRV-01 (Wazuh)', '#eef4f1', outline='#7a2e2e')
    bkp  = box(560,720,260,90,'Backup','BKP-SRV-01 (Veeam)', '#f6faf8')
    az   = box(860,720,260,90,'Azure Cloud','UAE North (me-central-1)', '#e8f0fb', outline='#0f6cbd')
    usr  = box(1180,720,240,90,'Users / Endpoints','320 staff · 3 sites', '#f6faf8')
    link(inet,fw); link(fw,dmz); link(fw,core); link(fw,vpn)
    link(core,srvz); link(core,idsvc); link(core,soc); link(srvz,bkp); link(srvz,az); link(vpn,usr); link(core,usr)
    dr.text((26,Hh-34), 'Fictional test diagram — %s. LEGACY-SRV-09 is NOT covered by backup (intentional gap).'%CO['short'], font=font(13), fill='#9aa4b2')
    img.save(path)

print('generating dataset at', ROOT)

# =================================================================== POLICIES
def std_sections(purpose, scope, statements, roles=None):
    roles = roles or ('The CISO owns this policy. IT and SOC teams implement controls. '
                      'All staff must comply. Compliance verifies adherence during internal audits.')
    return [
        ('1. Purpose', [purpose]),
        ('2. Scope', [scope + ' This policy applies to all employees, contractors, systems, and third parties of %s.' % CO['name_en']]),
        ('3. Policy Statements', ['%d.%d — %s' % (3, i+1, s) for i, s in enumerate(statements)]),
        ('4. Roles & Responsibilities', [roles]),
        ('5. Compliance & Enforcement', ['Non-compliance may result in disciplinary action. Controls are evidenced and reviewed at least annually and after major change, aligned to NCA ECC and ISO/IEC 27001.']),
        ('6. Exceptions', ['Exceptions require a formal request, risk assessment, and CISO approval, recorded in the Exception Register (EXC-*).']),
        ('7. Review', ['This document is reviewed annually or upon significant change by the CISO and approved by executive management.']),
    ]

POLICIES = [
 ('01_Information_Security_Policy.pdf','Information Security Policy','POL-ISP-001','3.2',
   std_sections('Establish the overarching framework to protect the confidentiality, integrity and availability of NDT information assets.',
     'All information assets, systems, and processing facilities.',
     ['NDT adopts a risk-based ISMS aligned to ISO/IEC 27001 and NCA ECC.',
      'Information is classified and handled per the Data Classification Policy.',
      'Access follows least-privilege and need-to-know.',
      'Security incidents are reported within 1 hour to the SOC.',
      'All third parties are assessed before access is granted.']), 60, 300),
 ('02_Cybersecurity_Governance_Policy.pdf','Cybersecurity Governance Policy','POL-GOV-002','2.1',
   std_sections('Define governance, roles, and oversight of the cybersecurity program.',
     'The enterprise cybersecurity governance model.',
     ['A Cybersecurity Steering Committee meets quarterly.',
      'The CISO reports cyber risk to executive management monthly.',
      'RACI is maintained for all key security controls.',
      'A cybersecurity strategy is reviewed annually.']), 80, 285),
 ('03_Access_Control_Policy.pdf','Access Control Policy','POL-AC-003','3.0',
   std_sections('Govern logical access to systems and data.',
     'All user, service, and privileged accounts.',
     ['Access is granted on least-privilege and approved by the data owner.',
      'User access is reviewed quarterly (see User_Access_Matrix).',
      'Privileged accounts require MFA and are logged.',
      'Dormant accounts (>45 days) are disabled.']), 70, 295),
 ('04_Identity_Management_Policy.pdf','Identity Management Policy','POL-IDM-004','2.4',
   std_sections('Manage the identity lifecycle from onboarding to offboarding.',
     'All identities in Microsoft Entra ID.',
     ['Identities are provisioned via HR-triggered workflow.',
      'Offboarding disables access within 24 hours.',
      'Conditional Access enforces device compliance and MFA.']), 90, 275),
 ('05_Password_Policy.pdf','Password Policy','POL-PWD-005','2.0',
   std_sections('Define password strength and lifecycle requirements.',
     'All authentication to NDT systems.',
     ['Minimum 12 characters with complexity.',
      'Passwords are never shared or reused across systems.',
      'Privileged passwords are vaulted and rotated every 90 days.']), 120, 245),
 ('06_MFA_Policy.pdf','Multi-Factor Authentication (MFA) Policy','POL-MFA-006','1.6',
   std_sections('Require MFA for access to NDT systems and cloud services.',
     'All remote access, cloud admin, and privileged access.',
     ['MFA is mandatory for ALL administrative and privileged accounts.',
      'MFA is enforced via Entra ID Conditional Access.',
      'Exceptions are prohibited for administrator accounts.']), 65, 300),  # policy says ALL admins -> contradicts Privileged_Accounts gap
 ('07_Asset_Management_Policy.pdf','Asset Management Policy','POL-AST-007','2.2',
   std_sections('Ensure all information assets are inventoried and owned.',
     'Hardware, software, cloud, and data assets.',
     ['Every asset has a unique code and an assigned owner.',
      'The asset inventory is reviewed quarterly.',
      'Assets are classified by criticality.']), 100, 265),
 ('08_Backup_Policy.pdf','Backup Policy','POL-BKP-008','2.3',
   std_sections('Ensure recoverability of critical data and systems.',
     'All production servers and critical data stores.',
     ['Critical systems are backed up DAILY with weekly full backups.',
      'Backups are tested (restore) quarterly.',
      'All production servers must be included in the backup scope.',
      'Backup success is monitored daily (see Backup_Success_Report).']), 75, 290),  # daily -> contradicts backup gap
 ('09_Business_Continuity_Plan.pdf','Business Continuity Plan (BCP)','POL-BCP-009','2.0',
   std_sections('Maintain business operations during disruption.',
     'Critical business services and their dependencies.',
     ['RTO for critical services is 4 hours; RPO is 24 hours.',
      'BCP is tested annually.',
      'Alternate site is Jeddah DR-2.']), 110, 255),
 ('10_Disaster_Recovery_Plan.pdf','Disaster Recovery Plan (DRP)','POL-DRP-010','2.0',
   std_sections('Restore IT services after a disaster.',
     'Data centers Riyadh DC-1 and Jeddah DR-2.',
     ['DR failover to Jeddah DR-2 within RTO.',
      'DR test performed annually with documented results.']), 115, 250),
 ('11_Incident_Response_Plan.pdf','Incident Response Plan','POL-IRP-011','2.5',
   std_sections('Detect, respond to, and recover from security incidents.',
     'All security events and incidents.',
     ['Incidents are triaged by the SOC within 1 hour.',
      'Severity is assigned and tracked in the Security_Incident_Register.',
      'Post-incident reviews are mandatory for High severity.',
      'Incidents must be formally closed with lessons learned.']), 55, 305),  # -> INC open gap
 ('12_Vulnerability_Management_Policy.pdf','Vulnerability Management Policy','POL-VUL-012','2.1',
   std_sections('Identify and remediate technical vulnerabilities.',
     'All servers, endpoints, and applications.',
     ['Authenticated scans run monthly.',
      'CRITICAL vulnerabilities are remediated within 15 days; HIGH within 30 days.',
      'Findings are tracked to closure (see Vulnerability report).']), 40, 320),  # -> High unremediated gap
 ('13_Patch_Management_Policy.pdf','Patch Management Policy','POL-PAT-013','2.0',
   std_sections('Keep systems current with security patches.',
     'All servers and endpoints.',
     ['Security patches are applied within 30 days of release.',
      'Patch compliance is reported monthly (see Patch_Compliance_Report).',
      'Legacy/unsupported systems require a documented exception.']), 85, 280),  # -> 92-day + legacy gap
 ('14_Secure_Development_Policy.pdf','Secure Software Development Policy','POL-SDL-014','1.4',
   std_sections('Embed security in the SDLC.',
     'All in-house developed applications.',
     ['Code is peer-reviewed and scanned (SAST) before release.',
      'Secrets are never committed to source control.',
      'Dependencies are scanned for known vulnerabilities.']), 130, 235),
 ('15_Change_Management_Policy.pdf','Change Management Policy','POL-CHG-015','2.2',
   std_sections('Control changes to production systems.',
     'All production changes.',
     ['Changes require a Change Request and CAB approval.',
      'Emergency changes are reviewed retrospectively.',
      'Rollback plans are documented.']), 95, 270),
 ('16_Third_Party_Security_Policy.pdf','Third Party Security Policy','POL-TPS-016','1.8',
   std_sections('Manage security risk from vendors and partners.',
     'All third parties with access to NDT systems/data.',
     ['Every vendor is security-assessed BEFORE onboarding.',
      'Contracts include security and data-protection clauses.',
      'Vendor risk is reviewed annually (see Vendor_Register).']), 105, 260),  # -> vendor no assessment gap
 ('17_Data_Classification_Policy.pdf','Data Classification Policy','POL-DCP-017','1.2',
   std_sections('Classify and handle information by sensitivity.',
     'All NDT information.',
     ['Data is classified as Public, Internal, Confidential, or Restricted.',
      'Restricted data is encrypted at rest and in transit.',
      'Handling rules are defined per classification.']), 400, -95),  # GAP: review overdue (next review in the PAST)
 ('18_Remote_Access_Policy.pdf','Remote Access Policy','POL-RMT-018','2.0',
   std_sections('Secure remote connectivity to NDT.',
     'All remote and VPN access.',
     ['Remote access requires VPN with MFA.',
      'Split tunneling is prohibited.',
      'Devices must be compliant (MDM).']), 88, 277),
 ('19_Email_Security_Policy.pdf','Email Security Policy','POL-EML-019','1.5',
   std_sections('Protect email against phishing and data loss.',
     'All corporate email.',
     ['DMARC, DKIM, and SPF are enforced.',
      'External emails are tagged.',
      'Suspicious emails are reported to the SOC.']), 92, 273),
 ('20_Acceptable_Use_Policy.pdf','Acceptable Use Policy','POL-AUP-020','2.1',
   std_sections('Define acceptable use of NDT IT resources.',
     'All users of NDT systems.',
     ['IT resources are for authorized business use.',
      'Users must not disable security controls.',
      'Violations are reported and may lead to disciplinary action.']), 78, 287),
]
for fn, title, did, ver, secs, ra, nr in POLICIES:
    build_pdf(os.path.join(ROOT,'Policies',fn), title, did, ver, secs, review_days_ago=ra, next_review_days=nr)
print('  policies:', len(POLICIES))

# =================================================================== REGISTERS
def R(path, sheet, headers, rows, note=''):
    build_xlsx(os.path.join(ROOT, path), sheet, headers, rows, note)

# Server Inventory (source of truth)
R('Registers/Server_Inventory.xlsx','Server Inventory',
  ['Server Code','Name','OS','Role','IP Address','Data Center','Owner','Criticality','In Backup','Last Patched'],
  [[s[0],s[1],s[2],s[3],s[4],s[5],s[6],s[7],
    ('No' if s[0]=='LEGACY-SRV-09' else 'Yes'),
    (d(92) if s[0]=='APP-SRV-02' else d(300) if s[0]=='LEGACY-SRV-09' else d(12))] for s in SERVERS],
  'NDT Server Inventory — interlinked with Patch, Backup, Vulnerability, Risk, Network.')

# Asset Inventory (note: LegacyBillingApp app is NOT here -> gap; LEGACY-SRV-09 has no owner)
R('Registers/Asset_Inventory.xlsx','Asset Inventory',
  ['Asset ID','Asset Name','Type','Owner','Location','Classification','Criticality'],
  [['AST-0001','APP-SRV-01','Server','App Team','Riyadh DC-1','Confidential','High'],
   ['AST-0002','APP-SRV-02','Server','App Team','Riyadh DC-1','Confidential','High'],
   ['AST-0003','DB-SRV-01','Server','DBA Team','Riyadh DC-1','Restricted','High'],
   ['AST-0004','WEB-SRV-01','Server','Web Team','DMZ','Confidential','High'],
   ['AST-0005','AD-SRV-01','Server','Infra Team','Riyadh DC-1','Restricted','High'],
   ['AST-0006','BKP-SRV-01','Server','Infra Team','Riyadh DC-1','Confidential','High'],
   ['AST-0007','SIEM-SRV-01','Server','SOC Team','Riyadh DC-1','Confidential','High'],
   ['AST-0008','FILE-SRV-01','Server','Infra Team','Riyadh DC-1','Confidential','Medium'],
   ['AST-0009','LEGACY-SRV-09','Server','(none)','Riyadh DC-1','Confidential','Medium'],
   ['AST-0010','DEV-LAP-217','Endpoint','(none)','Riyadh HQ','Internal','Low'],
   ['AST-0011','Entra ID Tenant','Cloud Identity','Infra Team','Azure','Restricted','High'],
   ['AST-0012','Azure Subscription','Cloud','Cloud Team','UAE North','Confidential','High']],
  'GAP: AST-0009 and AST-0010 have no owner. LegacyBillingApp (see Applications) is NOT listed as an asset.')

# Application Inventory
R('Registers/Application_Inventory.xlsx','Applications',
  ['App ID','Application','Owner','Hosting','Data Classification','MFA','Internet Facing'],
  [['APP-01','NDT FinTech Platform','App Team','APP-SRV-01/02','Restricted','Yes','Yes'],
   ['APP-02','Customer Portal','Web Team','WEB-SRV-01','Confidential','Yes','Yes'],
   ['APP-03','ERP (SAP)','Finance','Azure','Confidential','Yes','No'],
   ['APP-04','HR System','HR','Azure','Confidential','Yes','No'],
   ['APP-05','Internal Wiki','IT','FILE-SRV-01','Internal','No','No'],
   ['APP-06','LegacyBillingApp','(none)','LEGACY-SRV-09','Confidential','No','No']],
  'GAP: APP-06 LegacyBillingApp runs on LEGACY-SRV-09 but is NOT in the Asset Inventory, has no owner, no MFA.')

# Business Applications (criticality/BIA)
R('Registers/Business_Applications.xlsx','Business Apps',
  ['App ID','Application','Business Owner','RTO','RPO','Criticality'],
  [['APP-01','NDT FinTech Platform','Head of Product','4h','24h','Critical'],
   ['APP-02','Customer Portal','Head of Product','8h','24h','High'],
   ['APP-03','ERP (SAP)','CFO','8h','24h','High'],
   ['APP-04','HR System','HR Director','24h','24h','Medium'],
   ['APP-06','LegacyBillingApp','(none)','Undefined','Undefined','Medium']])

# Network Devices
R('Registers/Network_Devices.xlsx','Network Devices',
  ['Device ID','Device','Model','Mgmt IP','Location','Firmware','Last Review'],
  [['FW-01','Perimeter Firewall','FortiGate 600F','10.10.1.1','Riyadh DC-1','7.4.3',d(35)],
   ['SW-CORE-01','Core Switch 1','Cisco C9500','10.10.1.2','Riyadh DC-1','17.9',d(50)],
   ['SW-CORE-02','Core Switch 2','Cisco C9500','10.10.1.3','Riyadh DC-1','17.9',d(50)],
   ['VPN-01','VPN Gateway','FortiGate VPN','10.10.1.4','Riyadh DC-1','7.4.3',d(400)]],
  'GAP: VPN-01 last reviewed 400 days ago (review overdue).')

# User Access Matrix (Administrator MFA = No -> contradicts MFA Policy)
R('Registers/User_Access_Matrix.xlsx','User Access',
  ['User ID','Name','Department','Role','Systems','Privileged','MFA','Last Review'],
  [['U-001','Faisal Al-Harbi','Security','CISO','Entra, SIEM','Yes','Yes',d(80)],
   ['U-014','Administrator','IT','Domain Admin','AD, all servers','Yes','No',d(80)],  # GAP
   ['U-022','Noura Al-Qahtani','IT','IT Director','Entra, ERP, HR','Yes','Yes',d(80)],
   ['U-045','Yousef Al-Dossary','SOC','SOC Manager','SIEM, FW','Yes','Yes',d(80)],
   ['U-088','Sara Al-Otaibi','Finance','Accountant','ERP','No','Yes',d(80)],
   ['U-090','Sara Al-Otaibi','Finance','Accountant','ERP, HR, AD (admin)','Yes','Yes',d(80)],  # GAP: excessive privilege
   ['U-131','Khalid Al-Zahrani','Dev','Developer','App-SRV, DB-SRV (admin)','Yes','Yes',d(80)]],
  'GAP: Administrator (U-014) has MFA=No — contradicts MFA Policy. U-090 has excessive privileges (Finance user with AD admin).')

# Privileged Accounts
R('Registers/Privileged_Accounts.xlsx','Privileged Accounts',
  ['Account','Type','System','MFA','Vaulted','Last Rotated','Owner'],
  [['Administrator','Domain Admin','AD-SRV-01','No','No',d(210),'(shared)'],  # GAP
   ['sa','DB SysAdmin','DB-SRV-01','N/A','Yes',d(80),'DBA Team'],
   ['root','Root','WEB-SRV-01','N/A','Yes',d(60),'Web Team'],
   ['az-global-admin','Global Admin','Entra ID','Yes','Yes',d(45),'Infra Team'],
   ['fw-admin','Firewall Admin','FW-01','Yes','Yes',d(70),'SOC Team']],
  'GAP: built-in Administrator is shared, no MFA, no vault, rotated 210 days ago.')

# Risk Register (R-018 no treatment plan)
R('Registers/Risk_Register.xlsx','Risk Register',
  ['Risk ID','Description','Asset','Likelihood','Impact','Rating','Treatment Plan','Owner','Status'],
  [['R-001','Domain admin without MFA','AD-SRV-01','High','High','Critical','Enforce MFA for Administrator','CISO','Open'],
   ['R-005','Unpatched legacy server','LEGACY-SRV-09','High','Medium','High','Decommission or isolate','Infra Team','Open'],
   ['R-011','High vulns on public web','WEB-SRV-01','Medium','High','High','Patch within SLA','Web Team','In Treatment'],
   ['R-014','Vendor without security assessment','CloudPeak','Medium','Medium','Medium','Complete assessment','Compliance','Open'],
   ['R-018','Backup gap on legacy billing','LEGACY-SRV-09','Medium','High','High','(none)','(none)','Open']],  # GAP: no treatment
  'GAP: R-018 has no treatment plan and no owner.')

# Vendor Register (CloudPeak not assessed)
R('Registers/Vendor_Register.xlsx','Vendors',
  ['Vendor ID','Vendor','Service','Data Access','Security Assessment','Contract Security Clause','Last Review'],
  [['V-01','Azure (Microsoft)','Cloud IaaS/PaaS','Restricted','Completed','Yes',d(120)],
   ['V-02','Veeam','Backup software','Confidential','Completed','Yes',d(150)],
   ['V-03','CloudPeak Managed SOC','SOC co-managed','Restricted','NOT DONE','Pending',''],  # GAP
   ['V-04','FortiNet','Firewall/VPN','Internal','Completed','Yes',d(200)]],
  'GAP: V-03 CloudPeak has data access to Restricted data but NO security assessment (contradicts Third Party Policy).')

# Patch Compliance (APP-SRV-02 92 days; LEGACY-SRV-09 300 days)
R('Registers/Patch_Compliance_Report.xlsx','Patch Compliance',
  ['Server Code','OS','Last Patched','Days Since','Missing Critical','Missing High','Compliant (<=30d)'],
  [[s[0],s[2],
    (d(92) if s[0]=='APP-SRV-02' else d(300) if s[0]=='LEGACY-SRV-09' else d(12)),
    (92 if s[0]=='APP-SRV-02' else 300 if s[0]=='LEGACY-SRV-09' else 12),
    (1 if s[0]=='LEGACY-SRV-09' else 0),
    (2 if s[0] in ('APP-SRV-02','WEB-SRV-01') else 1 if s[0]=='LEGACY-SRV-09' else 0),
    ('No' if s[0] in ('APP-SRV-02','LEGACY-SRV-09') else 'Yes')] for s in SERVERS],
  'GAP: APP-SRV-02 not patched for 92 days; LEGACY-SRV-09 for 300 days (contradicts Patch Policy 30-day SLA).')

# Backup Success (DB-SRV-01 one failure; LEGACY-SRV-09 not covered)
R('Registers/Backup_Success_Report.xlsx','Backup Report',
  ['Job ID','Server Code','Schedule','Last Success','Last Result','Restore Tested'],
  [['BJ-01','APP-SRV-01','Daily',d(1),'Success',d(80)],
   ['BJ-02','APP-SRV-02','Daily',d(1),'Success',d(80)],
   ['BJ-03','DB-SRV-01','Daily',d(3),'FAILED (last run)',d(80)],  # GAP: recent failure
   ['BJ-04','WEB-SRV-01','Daily',d(1),'Success',d(80)],
   ['BJ-05','AD-SRV-01','Daily',d(1),'Success',d(80)],
   ['BJ-06','FILE-SRV-01','Daily',d(1),'Success',d(200)]],
  'GAP: DB-SRV-01 last backup FAILED. LEGACY-SRV-09 has NO backup job at all (contradicts Backup Policy).')

# Security Incident Register (INC open)
R('Registers/Security_Incident_Register.xlsx','Incidents',
  ['Incident ID','Date','Type','Severity','Affected Asset','Status','Closed Date'],
  [['INC-2026-009',d(120),'Phishing','Medium','U-088 mailbox','Closed',d(115)],
   ['INC-2026-012',d(60),'Malware','High','DEV-LAP-217','Closed',d(52)],
   ['INC-2026-014',d(18),'Suspicious admin login','High','AD-SRV-01','OPEN','']],  # GAP: open
  'GAP: INC-2026-014 (High, suspicious admin login on AD-SRV-01) is still OPEN — links to the no-MFA Administrator risk.')

# Awareness Training
R('Registers/Awareness_Training_Register.xlsx','Training',
  ['Employee','Department','Course','Completed','Score'],
  [['Faisal Al-Harbi','Security','Annual Security Awareness',d(90),'100%'],
   ['Noura Al-Qahtani','IT','Annual Security Awareness',d(90),'95%'],
   ['Sara Al-Otaibi','Finance','Phishing Awareness',d(90),'80%'],
   ['Khalid Al-Zahrani','Dev','Secure Coding',d(90),'88%'],
   ['(42 staff)','Various','Annual Security Awareness','Pending','—']],
  'Note: 42 staff have not completed annual awareness training (coverage gap).')

# Exception Register
R('Registers/Exception_Register.xlsx','Exceptions',
  ['Exception ID','Description','System','Approved By','Expiry','Status'],
  [['EXC-003','Legacy OS (2012 R2) retained','LEGACY-SRV-09','CISO',d(-30),'Expired'],  # GAP: expired
   ['EXC-007','Temporary local admin for migration','APP-SRV-02','IT Director',d(-10),'Expired']],
  'GAP: both exceptions are EXPIRED but the underlying conditions persist.')

# Third Party Assessment
R('Third_Party/Third_Party_Assessment_Register.xlsx','TP Assessment',
  ['Vendor ID','Vendor','Assessment Date','Score','Findings','Next Assessment'],
  [['V-01','Azure (Microsoft)',d(120),'A','0 High',d(-245)],
   ['V-02','Veeam',d(150),'B','1 Medium',d(-215)],
   ['V-03','CloudPeak Managed SOC','NOT ASSESSED','—','Unknown','Overdue']],
  'GAP: CloudPeak (V-03) never assessed despite Restricted data access.')

print('  registers: 15')

# =================================================================== REPORTS
def report(fn, title, did, sections, ra=20, cls='Confidential'):
    build_pdf(os.path.join(ROOT,'Reports',fn), title, did, '1.0', sections, review_days_ago=ra, next_review_days=90, classification=cls)

report('Quarterly_Vulnerability_Assessment_Report.pdf','Quarterly Vulnerability Assessment Report','RPT-VUL-Q2-2026',
  [('Executive Summary',['Authenticated scans covered 9 servers. 1 Critical and 5 High findings remain open beyond SLA. Public-facing WEB-SRV-01 carries 2 unremediated High vulnerabilities.']),
   ('Key Findings',['CRIT-001 — LEGACY-SRV-09 (Windows 2012 R2) end-of-support, 1 Critical unpatched. Open 300 days.',
                    'HIGH-014 — WEB-SRV-01 outdated OpenSSL (2 High CVEs). Open 34 days (SLA 30).',
                    'HIGH-021 — APP-SRV-02 missing OS security rollups (92 days).']),
   ('SLA Compliance',['Critical remediated within 15 days: 60%. High within 30 days: 55%. Below policy target.']),
   ('Recommendations',['Prioritize WEB-SRV-01 (internet-facing). Decommission/replace LEGACY-SRV-09. Restore patch cadence on APP-SRV-02.'])])

report('Executive_Cybersecurity_Report.pdf','Executive Cybersecurity Report — Q2 2026','RPT-EXE-Q2-2026',
  [('Overall Posture',['Program maturity is improving but 3 High risks remain open. MFA coverage is 92% but the built-in Administrator account is NOT covered — a critical exposure.']),
   ('Top Risks',['R-001 Domain admin without MFA (Critical).','R-011 High vulns on public web (High).','R-018 Backup gap on legacy billing (High, no treatment plan).']),
   ('Metrics',['Patch compliance 66%. Backup success 5/6 jobs (DB-SRV-01 last run failed). Awareness completion 87%. Open incidents: 1 (High).']),
   ('Executive Actions',['Approve budget to decommission legacy billing stack. Mandate MFA on all admin accounts within 14 days.'])])

report('Risk_Assessment_Report.pdf','Enterprise Risk Assessment Report','RPT-RISK-2026',
  [('Methodology',['Likelihood x Impact (5x5) per ISO 27005. 5 risks assessed, mapped to assets and controls.']),
   ('Risk Summary',['1 Critical, 3 High, 1 Medium. R-018 lacks a treatment plan and owner and must be assigned.']),
   ('Treatment',['R-001 and R-005 require immediate action. R-014 (vendor) pending assessment completion.'])])

report('Backup_Audit_Report.pdf','Backup Audit Report','RPT-BKP-2026',
  [('Scope',['6 backup jobs reviewed against the Backup Policy (daily, quarterly restore test).']),
   ('Findings',['DB-SRV-01: last scheduled backup FAILED — investigate job BJ-03.',
                'LEGACY-SRV-09: NOT included in any backup job — policy non-conformity.',
                'FILE-SRV-01: restore last tested 200 days ago (>quarterly).']),
   ('Conclusion',['Backup coverage is incomplete; policy requires all production servers to be backed up.'])])

report('Security_Awareness_Report.pdf','Security Awareness Program Report','RPT-AWR-2026',
  [('Summary',['Annual awareness completion is 87%. 42 staff pending. Phishing simulation click-rate 9%.']),
   ('Recommendation',['Escalate completion for the remaining 42 staff before quarter close.'])])

report('Internal_Audit_Report.pdf','Internal Audit Report — Cybersecurity Controls','RPT-IA-2026',
  [('Scope',['Access control, MFA, patch, backup, vendor management, incident management.']),
   ('Observations',['MFA policy not fully enforced (Administrator).','Patch SLA breached on 2 servers.','Vendor CloudPeak not assessed.','Data Classification Policy review overdue.','Incident INC-2026-014 open beyond target.']),
   ('Rating',['Partially Effective — corrective actions required.'])])

report('Compliance_Status_Report.pdf','NCA ECC Compliance Status Report','RPT-COMP-2026',
  [('Overview',['Self-assessed readiness against NCA ECC. Governance and asset domains strong; identity/access and vulnerability domains have open gaps.']),
   ('Gap Highlights',['1-2 Identity: Administrator without MFA.','2-x Vulnerability: High findings beyond SLA.','4-x Third Party: unassessed vendor.']),
   ('Note',['This is an internal readiness view — not an official certification or government accreditation.'])], cls='Confidential')
print('  reports: 7')

# =================================================================== SCREENSHOTS
S = lambda fn: os.path.join(ROOT,'Screenshots',fn)
png_dashboard(S('01_Entra_ID_Overview.png'),'Microsoft Entra ID','Identity Overview',
  [('Total users','327',True),('MFA registered','92%',False),('Guest users','14',True),
   ('Conditional Access','8 policies',True),('Risky sign-ins (7d)','3',False),('Admin roles','11',True)],'#0f6cbd',False)
png_dashboard(S('02_Conditional_Access.png'),'Microsoft Entra ID','Conditional Access',
  [('Policies','8',True),('Require MFA','Enabled',True),('Block legacy auth','Enabled',True),
   ('Device compliance','Required',True),('Admin exclusions','1 (Administrator)',False),('Report-only','2',True)],'#0f6cbd',False)
png_dashboard(S('03_MFA_Status.png'),'Microsoft Entra ID','Authentication Methods',
  [('MFA enforced (all)','92%',False),('Admins with MFA','10 / 11',False),('FIDO2 keys','23',True),
   ('SMS fallback','41',True),('No MFA (admin)','Administrator',False),('Passwordless','18%',True)],'#0f6cbd',False)
png_dashboard(S('04_Firewall_Dashboard.png'),'FortiGate FW-01','Firewall Dashboard',
  [('Threats blocked (24h)','1,204',True),('Active sessions','8,932',True),('Firmware','7.4.3',True),
   ('IPS','Enabled',True),('High alerts','2',False),('VPN tunnels','37',True)],'#b7602a',True)
png_dashboard(S('05_VPN_Configuration.png'),'FortiGate VPN-01','VPN Configuration',
  [('Type','SSL-VPN',True),('MFA','Required',True),('Split tunneling','Disabled',True),
   ('Active users','29',True),('Firmware review','400d ago',False),('Idle timeout','15 min',True)],'#b7602a',False)
png_dashboard(S('06_Backup_Jobs.png'),'Veeam Backup','Backup Jobs',
  [('Jobs','6',True),('Success (24h)','5 / 6',False),('Last full','2d ago',True),
   ('DB-SRV-01','FAILED',False),('LEGACY-SRV-09','No job',False),('Restore test','Quarterly',True)],'#1f8a5b',False)
png_dashboard(S('07_Patch_Dashboard.png'),'WSUS / Intune','Patch Dashboard',
  [('Compliant servers','5 / 9',False),('APP-SRV-02','92d overdue',False),('LEGACY-SRV-09','300d / EOL',False),
   ('Critical missing','1',False),('High missing','5',False),('Auto-approve','Enabled',True)],'#0f6cbd',False)
png_dashboard(S('08_Antivirus_Dashboard.png'),'Microsoft Defender','Endpoint Protection',
  [('Onboarded devices','318 / 320',True),('Threats (7d)','6 resolved',True),('Real-time','On',True),
   ('Out of date','2',False),('Exposure score','34 / 100',True),('AV signatures','Current',True)],'#0f6cbd',True)
png_dashboard(S('09_SIEM_Dashboard.png'),'Wazuh SIEM','SOC Dashboard',
  [('Events (24h)','2.1M',True),('Alerts (24h)','214',True),('Critical alerts','4',False),
   ('Open incidents','1',False),('Log sources','46',True),('Coverage','LEGACY-SRV-09 off',False)],'#7a2e2e',False)
png_dashboard(S('10_Email_Security.png'),'Microsoft Defender for O365','Email Security',
  [('DMARC','Enforced',True),('DKIM','Enabled',True),('SPF','Pass',True),
   ('Phishing blocked (7d)','487',True),('Quarantine','12',True),('Impersonation','On',True)],'#0f6cbd',True)
png_dashboard(S('11_DNS_Security.png'),'Azure DNS / Defender','DNS Security',
  [('Protected zones','4',True),('Blocked domains (24h)','93',True),('DNSSEC','Enabled',True),
   ('Tunneling alerts','0',True),('Filtering','On',True),('Logging','On',True)],'#0f6cbd',True)
png_dashboard(S('12_Cloud_Security.png'),'Microsoft Defender for Cloud','Cloud Security Posture',
  [('Secure score','62%',False),('Resources','214',True),('High recommendations','9',False),
   ('Region','UAE North',True),('Key Vault','Enabled',True),('Public storage','1 exposed',False)],'#0f6cbd',False)
png_dashboard(S('13_Azure_Resources.png'),'Microsoft Azure','Resource Inventory',
  [('Subscriptions','1',True),('VMs','12',True),('Storage accounts','7',True),
   ('Region','me-central-1',True),('NSGs','15',True),('Untagged resources','8',False)],'#0f6cbd',True)
png_dashboard(S('14_Windows_Server.png'),'Windows Server 2022','APP-SRV-01 System',
  [('Uptime','41 days',True),('CPU','38%',True),('Patched','12d ago',True),
   ('MFA (RDP)','Via VPN',True),('Defender','On',True),('Role','FinTech App',True)],'#2b5797',True)
png_dashboard(S('15_Linux_Server.png'),'Ubuntu 22.04 LTS','WEB-SRV-01 System',
  [('Uptime','88 days',True),('OpenSSL','Outdated',False),('High CVEs','2',False),
   ('Firewall (ufw)','Active',True),('SSH','Key-only',True),('Zone','DMZ',True)],'#772953',False)
print('  screenshots: 15')

# =================================================================== NETWORK + INFRA
png_network(os.path.join(ROOT,'Network','NDT_Network_Diagram.png'))
build_pdf(os.path.join(ROOT,'Infrastructure','Infrastructure_Overview.pdf'),'Infrastructure Overview','DOC-INF-001','1.0',
  [('Data Centers',['Primary: Riyadh DC-1. DR: Jeddah DR-2. Cloud: %s.'%CO['cloud']]),
   ('Core Systems',['Active Directory (AD-SRV-01), Entra ID (hybrid), SQL (DB-SRV-01), FinTech app (APP-SRV-01/02), Backup (BKP-SRV-01), SIEM (SIEM-SRV-01).']),
   ('Segmentation',['DMZ hosts WEB-SRV-01. Internal server zone segregated by VLAN. LEGACY-SRV-09 remains on the internal zone pending decommission.'])], review_days_ago=40, next_review_days=200)
print('  network + infrastructure: 2')

# ------- copy a few registers into their themed folders (Risk/Assets/Incidents/Training/Compliance) for the requested structure
import shutil
shutil.copy(os.path.join(ROOT,'Registers','Risk_Register.xlsx'), os.path.join(ROOT,'Risk','Risk_Register.xlsx'))
shutil.copy(os.path.join(ROOT,'Registers','Asset_Inventory.xlsx'), os.path.join(ROOT,'Assets','Asset_Inventory.xlsx'))
shutil.copy(os.path.join(ROOT,'Registers','Security_Incident_Register.xlsx'), os.path.join(ROOT,'Incidents','Security_Incident_Register.xlsx'))
shutil.copy(os.path.join(ROOT,'Registers','Awareness_Training_Register.xlsx'), os.path.join(ROOT,'Training','Awareness_Training_Register.xlsx'))
shutil.copy(os.path.join(ROOT,'Reports','Compliance_Status_Report.pdf'), os.path.join(ROOT,'Compliance','Compliance_Status_Report.pdf'))
# Evidence folder: primary machine-readable evidence for AI text extraction (xlsx/txt)
shutil.copy(os.path.join(ROOT,'Registers','User_Access_Matrix.xlsx'), os.path.join(ROOT,'Evidence','User_Access_Matrix.xlsx'))
shutil.copy(os.path.join(ROOT,'Registers','Patch_Compliance_Report.xlsx'), os.path.join(ROOT,'Evidence','Patch_Compliance_Report.xlsx'))
shutil.copy(os.path.join(ROOT,'Registers','Backup_Success_Report.xlsx'), os.path.join(ROOT,'Evidence','Backup_Success_Report.xlsx'))

# =================================================================== README
readme_sections = [
 ('Company Overview', [
   '%s («%s») is a FICTIONAL Saudi mid-to-large technology company created solely to test the 1SaudiCyber compliance platform. CR %s, VAT %s, HQ %s, branches in %s.' % (CO['name_en'], CO['name_ar'], CO['cr'], CO['vat'], CO['hq'], ', '.join(CO['branches'])),
   'Sector: %s. Employees: %d. Cloud: %s. Identity: %s. Data centers: %s (primary) and %s (DR).' % (CO['sector'], CO['employees'], CO['cloud'], CO['idp'], CO['dc_primary'], CO['dc_dr']),
   'Key people (fictional): %s (CEO), %s (CISO), %s (IT Director), %s (SOC Manager), %s (Compliance).' % (CO['ceo'], CO['ciso'], CO['itd'], CO['socm'], CO['cco']),
 ]),
 ('Package Structure', [
   'Policies/ — 20 governance & security policies (PDF).',
   'Registers/ — 15 Excel registers (assets, users, risks, patch, backup, incidents, vendors...).',
   'Reports/ — 7 assessment/audit reports (PDF).',
   'Screenshots/ — 15 evidence screenshots (Entra ID, MFA, Firewall, Backup, SIEM, Cloud...).',
   'Network/ — high-level network architecture diagram (PNG).',
   'Infrastructure/, Risk/, Assets/, Incidents/, Training/, Third_Party/, Compliance/ — themed copies.',
   'Evidence/ — machine-readable evidence (XLSX) recommended for AI text-extraction testing.',
 ]),
 ('Interlinked Data (single source of truth: servers)', [
   'Each server (e.g. APP-SRV-01, DB-SRV-01, WEB-SRV-01, LEGACY-SRV-09) appears consistently across Server_Inventory, Patch_Compliance, Backup_Success, the Vulnerability report, Risk_Register, and the Network diagram.',
   'IPs, owners, and criticality are consistent across all files.',
 ]),
 ('Deliberate, Realistic Gaps (for gap-discovery / RFI / AI advisory testing)', [
   '1. Administrator (U-014) has MFA=No — CONTRADICTS the MFA Policy («all admins»).',
   '2. APP-SRV-02 not patched for 92 days; LEGACY-SRV-09 for 300 days — breaches Patch Policy (30-day SLA).',
   '3. WEB-SRV-01 has 2 unremediated HIGH vulnerabilities (public-facing).',
   '4. Data Classification Policy review is OVERDUE (next-review date in the past).',
   '5. Vendor CloudPeak (V-03) has Restricted data access but NO security assessment.',
   '6. Assets AST-0009 (LEGACY-SRV-09) and AST-0010 have no owner.',
   '7. User U-090 (Finance) has excessive privileges (AD admin).',
   '8. DB-SRV-01 last backup FAILED; LEGACY-SRV-09 has NO backup job.',
   '9. Incident INC-2026-014 (High) is still OPEN.',
   '10. Risk R-018 has no treatment plan and no owner.',
   '11. LEGACY-SRV-09 is in the network but not in Backup scope.',
   '12. LegacyBillingApp (APP-06) is not in the Asset Inventory.',
   'Cross-links: gap #1 (no-MFA admin) ↔ incident #9 (suspicious admin login on AD-SRV-01) ↔ risk R-001.',
 ]),
 ('How to Test 1SaudiCyber', [
   'Upload XLSX registers (Evidence/) to test extraction, table parsing, classification, and AI advisory (v2 consultative engine).',
   'Upload PDF policies/reports to test the classification + «needs human review» (OCR-deferred) path.',
   'Map evidence to controls, run gap analysis, and confirm the platform surfaces the 12 gaps above.',
   'Create RFIs where evidence is insufficient (e.g. request the missing Backup Success proof for LEGACY-SRV-09).',
   'The AI is advisory ONLY — the human auditor issues the final verdict.',
 ]),
 ('Notice', [
   'All names, numbers, and data are FICTIONAL. No real company, person, secret, or credential is used. For platform testing/training only.',
 ]),
]
build_pdf(os.path.join(ROOT,'README.pdf'), 'Enterprise Test Dataset — README', 'DOC-README-001', '1.0',
          readme_sections, review_days_ago=0, next_review_days=365, classification='Internal', frameworks='NCA ECC, ISO 27001, NIST CSF')
print('  README.pdf')


print('DONE')
