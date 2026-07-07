"""
Report generation (FR-011 / FR-007.12 / FR-009.13):
PDF reports via reportlab, Excel export via openpyxl, and audit certificates.
All functions return bytes so views can serve them as downloads.
"""
import io
from datetime import date

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle)

GOV_GREEN = colors.HexColor('#1a4731')

# R5 — every generated document must state, verbatim, that it is NOT an official
# certification / government accreditation and does not represent NCA / Aramco / SABIC.
DISCLAIMER_EN = ('This document is an internal readiness assessment produced by CyberTrust KSA. '
                 'It is NOT an official cybersecurity certification or government accreditation, '
                 'and does not represent NCA, Aramco, or SABIC.')
DISCLAIMER_AR = ('هذا المستند تقييم جاهزية داخلي من CyberTrust KSA، وليس شهادة امتثال رسمية أو '
                 'اعتمادًا حكوميًا، ولا يمثّل الهيئة الوطنية للأمن السيبراني أو أرامكو أو سابك.')


def _styles():
    ss = getSampleStyleSheet()
    ss.add(ParagraphStyle('H1x', parent=ss['Heading1'], textColor=GOV_GREEN))
    ss.add(ParagraphStyle('Smallx', parent=ss['Normal'], fontSize=8, textColor=colors.grey))
    return ss


def _disclaimer_elements(ss):
    """Bilingual legal disclaimer block appended to every report/PDF."""
    return [
        Spacer(1, 8 * mm),
        Paragraph(DISCLAIMER_EN, ss['Smallx']),
        Paragraph(DISCLAIMER_AR, ss['Smallx']),
    ]


def gap_analysis_pdf(company):
    """Bilingual gap-analysis / compliance summary PDF for a company."""
    from compliance.models import CompanyControl, Framework
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=f'{company.name} — Gap Analysis')
    ss = _styles()
    elements = [
        Paragraph('CyberTrust KSA — Gap Analysis Report', ss['H1x']),
        Paragraph(f'{company.name} ({company.cr_number})', ss['Normal']),
        Paragraph(f'Generated: {date.today():%Y-%m-%d}', ss['Smallx']),
        Spacer(1, 8 * mm),
    ]
    rows = [['Framework', 'Total', 'Compliant', 'Non-Compliant', 'Partial', 'Score %']]
    qs = CompanyControl.objects.filter(company=company)
    for fw in Framework.objects.filter(is_active=True):
        fwq = qs.filter(control__framework=fw)
        total = fwq.count()
        if not total:
            continue
        comp = fwq.filter(status='compliant').count()
        nonc = fwq.filter(status='non_compliant').count()
        part = fwq.filter(status='partially_compliant').count()
        rows.append([fw.code, total, comp, nonc, part, f'{comp/total*100:.1f}'])
    table = Table(rows, hAlign='LEFT')
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), GOV_GREEN),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 6 * mm))
    elements.append(Paragraph(f'Overall compliance: {company.overall_compliance_score:.1f}%', ss['Normal']))
    elements.append(Paragraph('تقرير تحليل الفجوات — ملخّص الامتثال عبر الأطر الثلاثة.', ss['Normal']))
    elements.extend(_disclaimer_elements(ss))
    doc.build(elements)
    return buf.getvalue()


def compliance_excel(company):
    """Excel export of every company control with its status (FR-011.9 / FR-004.7)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from compliance.models import CompanyControl
    wb = Workbook()
    ws = wb.active
    ws.title = 'Controls'
    headers = ['Framework', 'Control ID', 'Title', 'Domain', 'Priority', 'Status', 'AI Verdict', 'Confidence']
    ws.append(headers)
    for i in range(1, len(headers) + 1):
        c = ws.cell(1, i)
        c.font = Font(bold=True, color='FFFFFF')
        c.fill = PatternFill('solid', fgColor='1A4731')
    for cc in CompanyControl.objects.filter(company=company).select_related(
            'control', 'control__framework', 'control__domain'):
        ws.append([
            cc.control.framework.code, cc.control.control_id, cc.control.title[:80],
            cc.control.domain.name[:40], cc.control.priority, cc.status,
            cc.ai_verdict, cc.ai_confidence,
        ])
    for col, w in zip('ABCDEFGH', [16, 12, 50, 28, 10, 18, 18, 10]):
        ws.column_dimensions[col].width = w
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def certificate_pdf(company, framework_code, certificate_number, expiry):
    """Internal readiness ACKNOWLEDGEMENT PDF — NOT an official certificate.

    R5: the platform is not a certification body, so this document never claims official
    compliance. It records that an internal readiness review was completed and carries the
    same non-certification disclaimer as every other report.
    """
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title='Cybersecurity Readiness Acknowledgement')
    ss = _styles()
    center = ParagraphStyle('C', parent=ss['Title'], alignment=1, textColor=GOV_GREEN)
    elements = [
        Spacer(1, 30 * mm),
        Paragraph('Cybersecurity Readiness Acknowledgement', center),
        Spacer(1, 6 * mm),
        Paragraph('إقرار جاهزية للأمن السيبراني (ليس شهادة رسمية)',
                  ParagraphStyle('Cc', parent=ss['Title'], alignment=1)),
        Spacer(1, 14 * mm),
        Paragraph(f'This acknowledges that <b>{company.name}</b> (CR {company.cr_number})',
                  ParagraphStyle('B', parent=ss['Normal'], alignment=1, fontSize=12)),
        Paragraph(f'completed an internal readiness review for <b>{framework_code}</b>.',
                  ParagraphStyle('B2', parent=ss['Normal'], alignment=1, fontSize=12)),
        Spacer(1, 10 * mm),
        Paragraph(f'Reference No: {certificate_number}', ParagraphStyle('m', parent=ss['Normal'], alignment=1)),
        Paragraph(f'Issued: {date.today():%Y-%m-%d}   ·   Review window until: {expiry:%Y-%m-%d}',
                  ParagraphStyle('m2', parent=ss['Normal'], alignment=1)),
    ]
    elements.extend(_disclaimer_elements(ss))
    doc.build(elements)
    return buf.getvalue()
