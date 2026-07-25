"""
Phase 3F — advisory AI/OCR evidence analysis pipeline.

Analyzes an EvidenceSubmission file and stores a DRAFT EvidenceAnalysisResult.
AI is an assistant ONLY: it never decides compliant/non-compliant, never accepts/
rejects evidence, never creates ControlAssessment. The final decision is the
auditor's (Phase 3G).

Safety:
  * text extraction is size/type limited; full file content is never logged.
  * heavy OCR (pdf/images) is intentionally NOT run here -> needs_human_review.
  * OpenAI key is read from settings/env only; missing key fails gracefully.
  * tenant scoping is the caller's responsibility (commands/views filter by company).

Standalone module (not a `services/` package) to avoid colliding with compliance/services.py.
"""
import hashlib
import json

from django.conf import settings
from django.utils import timezone

from compliance.models import EvidenceSubmission, EvidenceAnalysisResult

PROMPT_VERSION = 'advisory-v2-consultative'


def _content_fingerprint(control, requirement, text, related_context):
    """Stable hash of everything that feeds the AI call. If it matches a prior
    completed analysis for the same submission, the (expensive) model call is skipped."""
    h = hashlib.sha256()
    for part in (PROMPT_VERSION, str(getattr(control, 'id', '')),
                 str(getattr(requirement, 'id', '')), text or '', related_context or ''):
        h.update(part.encode('utf-8', 'ignore'))
        h.update(b'\x00')
    return h.hexdigest()
MAX_EXTRACT_CHARS = 20000   # cap stored/extracted text
MAX_XLSX_ROWS = 500
TEXT_LIKE = {'txt', 'csv'}
# pdf/images deliberately deferred (no heavy OCR in this phase).
OCR_DEFERRED = {'pdf', 'png', 'jpg', 'jpeg'}

# Consultative evidence-analysis engine (advisory ONLY). Encodes a 15-phase
# senior-consultant methodology (intake → classification → mapping → extraction →
# quality → control analysis → cross-referencing → contradictions → completeness →
# risk → remediation → confidence → gaps → RFI). The AI is a decision-support engine
# — it NEVER issues a compliance verdict; the human auditor is authoritative.
ADVISORY_SYSTEM_PROMPT = (
    "You are the Cybersecurity Compliance Intelligence Engine inside 1SaudiCyber — a senior "
    "GRC consultant (20+ yrs: NCA ECC/CSCC, ISO 27001, NIST CSF, CIS, SAMA, cloud security, "
    "internal/external audit). You provide ADVISORY analysis to help a company and a human "
    "auditor decide — you are NOT an accreditation body.\n"
    "HARD RULES (never violate): you DO NOT issue a final compliance decision; DO NOT mark a "
    "control compliant/non-compliant; DO NOT accept/reject evidence; DO NOT change any status; "
    "DO NOT close an RFI. The human auditor's professional verdict is final. Never invent evidence "
    "not present in the text; if text is empty/insufficient, say so and set needs_human_review=true.\n"
    "METHOD: think like a consultant AND an auditor AND a risk manager. Do NOT rely on one file — "
    "cross-reference the CURRENT evidence with the OTHER evidence already analyzed for this control "
    "(provided), look for contradictions between policy and implementation, and note missing links. "
    "Every conclusion must state the evidence it rests on and whether it is confirmed or needs "
    "further verification. Prefer 'evidence suggests a possible gap requiring verification' over "
    "absolute compliance statements.\n"
    "Return ONLY a JSON object with these keys:\n"
    "  document_classification: {type, confidence}  // Policy/Procedure/Log/Config/Screenshot/Report/"
    "Certificate/Register/RiskRegister/IncidentReport/BackupReport/NetworkDiagram/AuditReport/Contract/"
    "TrainingRecord/VendorAssessment/Other\n"
    "  summary: string  // concise executive summary\n"
    "  requirement_match: string  // how the evidence relates to the evidence requirement\n"
    "  extracted_entities: {assets, servers, users, mfa, backup, firewall, encryption, cloud_services, "
    "vendors, ip_addresses, urls, dates, versions}  // each an array of short strings actually found\n"
    "  quality: {grade: Excellent|Good|Acceptable|Weak|Insufficient, reason}\n"
    "  control_analysis: string  // consultative reasoning, e.g. 'policy found but no proof of "
    "implementation'; connect the evidence pieces\n"
    "  contradictions: [string]  // policy-vs-implementation or evidence-vs-evidence conflicts\n"
    "  completeness: {completeness, consistency, freshness, traceability, authenticity, coverage}  // each 0-100\n"
    "  potential_gaps: string\n"
    "  risk_flags: [string]  // short flags\n"
    "  risk_analysis: [string]  // gap -> plausible risk (e.g. no MFA -> credential theft/privilege escalation)\n"
    "  remediation_suggestions: [{priority: high|medium|low, action}]  // ordered by priority; suggestions only\n"
    "  missing_evidence: [string]  // what is missing to reach a stronger conclusion\n"
    "  rfi_suggestions: [string]  // concrete info to request from the company (do NOT create/close RFIs)\n"
    "  confidence: 0.0-1.0  // based on: amount + quality + coherence + freshness of evidence, and # of contradictions\n"
    "  confidence_rationale: string  // WHY this score\n"
    "  needs_human_review: boolean\n"
    "All prose in Arabic. Keep arrays short and grounded strictly in the provided text."
)


def extract_text_from_submission(submission):
    """Return (text, truncated, note). Size/type limited; no heavy OCR; no full-content logging."""
    ext = (submission.file_type or '').lower()
    try:
        if ext in TEXT_LIKE:
            with submission.uploaded_file.open('rb') as f:
                raw = f.read(MAX_EXTRACT_CHARS + 1)
            text = raw.decode('utf-8', errors='replace')
            truncated = len(raw) > MAX_EXTRACT_CHARS
            return text[:MAX_EXTRACT_CHARS], truncated, ''
        if ext == 'docx':
            try:
                from docx import Document
            except Exception:
                return '', False, 'python-docx not available'
            doc = Document(submission.uploaded_file.open('rb'))
            parts = [p.text for p in doc.paragraphs if p.text.strip()]
            text = '\n'.join(parts)
            return text[:MAX_EXTRACT_CHARS], len(text) > MAX_EXTRACT_CHARS, ''
        if ext == 'xlsx':
            try:
                from openpyxl import load_workbook
            except Exception:
                return '', False, 'openpyxl not available'
            wb = load_workbook(submission.uploaded_file.open('rb'), read_only=True, data_only=True)
            parts, rows = [], 0
            for ws in wb.worksheets:
                for row in ws.iter_rows(values_only=True):
                    cells = [str(c) for c in row if c is not None]
                    if cells:
                        parts.append('\t'.join(cells)); rows += 1
                    if rows >= MAX_XLSX_ROWS:
                        break
                if rows >= MAX_XLSX_ROWS:
                    break
            wb.close()
            text = '\n'.join(parts)
            return text[:MAX_EXTRACT_CHARS], len(text) > MAX_EXTRACT_CHARS, ''
        if ext in OCR_DEFERRED:
            # OCR the image/PDF via the platform's tesseract pipeline (ara+eng). Heavy —
            # run async in production (EVIDENCE_ASYNC_ENABLED) so it stays off the request thread.
            try:
                from ai_engine.services import extract_text_from_image, extract_text_from_pdf
                path = submission.uploaded_file.path  # local storage; S3 has no local path
                res = extract_text_from_pdf(path) if ext == 'pdf' else extract_text_from_image(path)
                txt = (res.get('text', '') if isinstance(res, dict) else str(res or '')) or ''
                conf = res.get('confidence') if isinstance(res, dict) else None
                if txt.strip():
                    note = 'ocr' + (' (confidence %.2f)' % conf if isinstance(conf, (int, float)) else '')
                    return txt[:MAX_EXTRACT_CHARS], len(txt) > MAX_EXTRACT_CHARS, note
                return '', False, 'OCR produced no readable text; needs human review'
            except Exception as exc:
                return '', False, 'OCR error: %s; needs human review' % type(exc).__name__
        return '', False, f'unsupported file type: {ext}'
    except Exception as exc:  # never leak file content; only the error class/message
        return '', False, f'extraction error: {type(exc).__name__}'


def _run_ai(control, requirement, text, related_context=''):
    """Call the AI provider for an advisory analysis. Returns (result_dict, error, model, provider).

    P0-04 — FAIL-CLOSED on data residency: the CENTRAL policy (external_ai_allowed), not mere API
    key presence, decides. When external AI is not allowed, this returns a safe 'not allowed'
    result WITHOUT building a client, constructing a prompt, or sending any tenant evidence text —
    the caller then records a needs-human-review analysis (never a compliance decision).
    """
    from ai_engine.services import (external_ai_allowed, ai_enabled, get_openai_client,
                                     ExternalAINotAllowed)
    if not external_ai_allowed():
        # Distinguish the two fail-closed reasons (mirrors ai_engine.ai_advisory_state) so the
        # recorded message is accurate: a missing key is reported as such; otherwise the residency
        # policy is the blocker. Either way NO client is built and NO evidence text is sent.
        if not ai_enabled():
            return None, 'AI provider not configured (no OPENAI_API_KEY)', '', ''
        return None, 'external AI disabled by data-residency policy', '', ''
    try:
        client = get_openai_client()
        model = getattr(settings, 'OPENAI_MODEL', 'gpt-4o')
        related_block = (f"\nOTHER EVIDENCE ALREADY ANALYZED FOR THIS CONTROL (cross-reference against "
                         f"these; look for contradictions and coverage):\n---\n{related_context[:3000]}\n---\n"
                         if related_context else "\n(No other evidence has been analyzed for this control yet.)\n")
        user = (f"CONTROL: {control.control_id} — {control.title}\n{control.description}\n\n"
                f"EVIDENCE REQUIREMENT: {getattr(requirement, 'title', '')} — "
                f"{getattr(requirement, 'description', '')}\n\n"
                f"CURRENT EVIDENCE TEXT:\n---\n{text[:8000]}\n---\n"
                f"{related_block}\n"
                "Return ONLY the JSON object described. The final compliance decision is the auditor's.")
        resp = client.chat.completions.create(
            model=model, temperature=0.2, response_format={'type': 'json_object'},
            messages=[{'role': 'system', 'content': ADVISORY_SYSTEM_PROMPT},
                      {'role': 'user', 'content': user}])
        data = json.loads(resp.choices[0].message.content)
        return data, '', model, 'openai'
    except ExternalAINotAllowed:
        # Defense-in-depth: the provider-boundary tripwire fired (residency flipped/forgotten guard).
        return None, 'external AI disabled by data-residency policy', '', ''
    except Exception as exc:
        return None, f'AI error: {type(exc).__name__}', '', 'openai'


def analyze_evidence_submission(submission, *, apply=False):
    """Analyze one submission and (optionally) persist an EvidenceAnalysisResult. Idempotent.

    Never creates ControlAssessment/CompanyControl; never accepts/rejects evidence.
    """
    item = submission.checklist_item
    requirement = item.evidence_requirement
    control = requirement.control
    text, truncated, note = extract_text_from_submission(submission)

    # Decide status/content without making any compliance decision.
    fields = dict(
        company=submission.company, checklist_item=item, control=control,
        evidence_requirement=requirement, extracted_text=text,
        extracted_text_truncated=truncated, prompt_version=PROMPT_VERSION,
        risk_flags=[], analysis_metadata={'extraction_note': note})

    if not text:
        fields.update(status='needs_human_review',
                      error_message=note or 'No text extracted', summary='', confidence=None)
    else:
        # Cross-reference context: prior advisory analyses for the SAME control (never one file alone).
        related_context = ''
        try:
            prior = (EvidenceAnalysisResult.objects
                     .filter(company=submission.company, control=control)
                     .exclude(evidence_submission=submission).order_by('-id')[:5])
            related_context = '\n'.join(
                f"- {(p.summary or '')[:300]}" for p in prior if (p.summary or '').strip())
        except Exception:
            related_context = ''
        # Cache: if this submission already has a completed analysis whose inputs
        # (text + control + requirement + prompt version + related context) are unchanged,
        # reuse it instead of paying for another model call. Any input change invalidates it.
        fingerprint = _content_fingerprint(control, requirement, text, related_context)
        if apply:
            existing = EvidenceAnalysisResult.objects.filter(evidence_submission=submission).first()
            if (existing is not None and existing.status == 'completed'
                    and (existing.analysis_metadata or {}).get('content_hash') == fingerprint):
                return {'submission': submission.id, 'status': existing.status,
                        'id': existing.id, 'cached': True}
        data, err, model, provider = _run_ai(control, requirement, text, related_context)
        if data is None:
            fields.update(status='needs_human_review', error_message=err, model_used=model,
                          provider=provider, summary='', confidence=None)
        else:
            # Top-level fields keep the existing UI contract; the full 15-phase consultative
            # output is preserved (additively) in analysis_metadata['advisory'].
            _adv_keys = ('document_classification', 'extracted_entities', 'quality',
                         'control_analysis', 'contradictions', 'completeness', 'risk_analysis',
                         'remediation_suggestions', 'missing_evidence', 'rfi_suggestions',
                         'confidence_rationale', 'needs_human_review')
            advisory = {k: data.get(k) for k in _adv_keys if data.get(k) is not None}
            meta = dict(fields.get('analysis_metadata') or {})
            meta['advisory'] = advisory
            meta['cross_referenced'] = bool(related_context)
            meta['content_hash'] = fingerprint
            fields.update(
                status='completed', model_used=model, provider=provider,
                summary=str(data.get('summary', ''))[:4000],
                requirement_match=str(data.get('requirement_match', ''))[:4000],
                potential_gaps=str(data.get('potential_gaps', ''))[:4000],
                risk_flags=data.get('risk_flags', []) if isinstance(data.get('risk_flags'), list) else [],
                confidence=data.get('confidence') if isinstance(data.get('confidence'), (int, float)) else None,
                analysis_metadata=meta, error_message='')

    if not apply:
        return {'submission': submission.id, 'would_status': fields['status'],
                'extracted_chars': len(text)}

    result, _ = EvidenceAnalysisResult.objects.update_or_create(
        evidence_submission=submission, defaults=fields)
    return {'submission': submission.id, 'status': result.status, 'id': result.id}


def batch_analyze_pending_submissions(company=None, *, apply=False):
    """Analyze submissions that have no completed analysis yet (optionally one company)."""
    qs = EvidenceSubmission.objects.all()
    if company is not None:
        qs = qs.filter(company=company)
    qs = qs.exclude(analysis__status='completed')
    results = [analyze_evidence_submission(s, apply=apply) for s in qs]
    return {'analyzed': len(results), 'results': results}
