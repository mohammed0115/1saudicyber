"""
Management command to seed the FULL 334-control set for CyberTrust KSA
from the authoritative source spreadsheets shipped in compliance/data/:

  - CyberTrust_KSA_Consolidated_Compliance_Rules_v2.xlsx  (control statements)
  - CyberTrust_KSA_Evidence_Matrix.xlsx                   (evidence type + Arabic guidance)

Usage:
    python manage.py seed_all_controls          # wipes the 3 frameworks' controls and reseeds (default)
    python manage.py seed_all_controls --keep   # upsert without wiping

This closes the gap where only 197 of the required 334 controls were seeded.
"""
import re
from pathlib import Path

from django.core.management.base import BaseCommand
from django.db import transaction
from openpyxl import load_workbook

from compliance.models import Framework, Domain, Control

DATA_DIR = Path(__file__).resolve().parents[2] / 'data'
RULES_FILE = DATA_DIR / 'CyberTrust_KSA_Consolidated_Compliance_Rules_v2.xlsx'
EVIDENCE_FILE = DATA_DIR / 'CyberTrust_KSA_Evidence_Matrix.xlsx'

# Source label (in the spreadsheets) -> Framework definition
FRAMEWORKS = {
    'NCA ECC:2018': {
        'code': 'NCA_ECC',
        'name': 'NCA Essential Cybersecurity Controls',
        'name_ar': 'الضوابط الأساسية للأمن السيبراني',
        'version': '2.0',
        'description': 'National Cybersecurity Authority Essential Cybersecurity Controls.',
    },
    'Aramco SACS-002': {
        'code': 'ARAMCO_SACS002',
        'name': 'Saudi Aramco SACS-002',
        'name_ar': 'معيار أرامكو للأمن السيبراني',
        'version': '5.0',
        'description': 'Saudi Aramco Third-Party Cybersecurity Standard for suppliers and contractors.',
    },
    'SABIC CyberTrust': {
        'code': 'SABIC_CT',
        'name': 'SABIC CyberTrust',
        'name_ar': 'معيار سابك للثقة السيبرانية',
        'version': '1.0',
        'description': 'SABIC CyberTrust Standard for supplier cybersecurity compliance.',
    },
}

# Evidence-type description (from the Evidence Matrix) -> model EVIDENCE_TYPE_CHOICES key.
# Order matters: first keyword match wins.
EVIDENCE_KEYWORDS = [
    ('screenshot', 'screenshot'),
    ('configuration', 'config'),
    ('config', 'config'),
    ('scan report', 'report'),
    ('report', 'report'),
    ('log', 'log'),
    ('procedure', 'procedure'),
    ('policy', 'policy'),
    ('contract', 'other'),
    ('nda', 'other'),
    ('training', 'other'),
    ('attendance', 'other'),
    ('inventory', 'other'),
    ('interview', 'interview'),
    ('certificate', 'certificate'),
]

# Theme (from the Evidence Matrix) -> priority. Default is 'medium'.
THEME_PRIORITY = {
    'governance & policy': 'high',
    'identity & access management': 'high',
    'incident management': 'high',
    'vulnerability management': 'high',
    'data protection': 'high',
    'network security': 'high',
}


def _clean(value):
    if value is None:
        return ''
    return re.sub(r'\s+', ' ', str(value)).strip()


def _evidence_type(desc):
    low = desc.lower()
    for keyword, choice in EVIDENCE_KEYWORDS:
        if keyword in low:
            return choice
    return 'policy'


def _priority(theme):
    return THEME_PRIORITY.get(theme.lower().strip(), 'medium')


def _load_evidence_index():
    """Build {(framework_code, control_id): {type, theme, guidance_ar}} from the Evidence Matrix."""
    index = {}
    if not EVIDENCE_FILE.exists():
        return index
    wb = load_workbook(EVIDENCE_FILE, read_only=True, data_only=True)
    ws = wb['Evidence Matrix']
    rows = ws.iter_rows(values_only=True)
    next(rows, None)  # header: Control ID, Source, Statement, Themes, Evidence Type, Evidence Requirements (Arabic)
    for r in rows:
        if not r or not r[0]:
            continue
        cid = _clean(r[0])
        src = _clean(r[1])
        fw = FRAMEWORKS.get(src)
        if not fw:
            continue
        index[(fw['code'], cid)] = {
            'theme': _clean(r[3]),
            'evidence_desc': _clean(r[4]),
            'guidance_ar': _clean(r[5]),
        }
    wb.close()
    return index


class Command(BaseCommand):
    help = 'Seed the full 334-control set from the consolidated source spreadsheets.'

    def add_arguments(self, parser):
        parser.add_argument(
            '--keep', action='store_true',
            help='Upsert without wiping existing controls for the 3 frameworks.',
        )

    @transaction.atomic
    def handle(self, *args, **options):
        if not RULES_FILE.exists():
            self.stderr.write(self.style.ERROR(f'Missing data file: {RULES_FILE}'))
            return

        evidence = _load_evidence_index()

        # 1) Frameworks
        fw_objs = {}
        for src, meta in FRAMEWORKS.items():
            obj, _ = Framework.objects.update_or_create(
                code=meta['code'],
                defaults={
                    'name': meta['name'], 'name_ar': meta['name_ar'],
                    'version': meta['version'], 'description': meta['description'],
                    'is_active': True,
                },
            )
            fw_objs[meta['code']] = obj

        # 2) Optional wipe (default) — cascades to controls/domains for a clean 334.
        if not options['keep']:
            codes = [m['code'] for m in FRAMEWORKS.values()]
            Control.objects.filter(framework__code__in=codes).delete()
            Domain.objects.filter(framework__code__in=codes).delete()

        # 3) Read consolidated rules and create domains + controls
        wb = load_workbook(RULES_FILE, read_only=True, data_only=True)
        ws = wb['Consolidated_Rules']
        rows = ws.iter_rows(values_only=True)
        next(rows, None)  # blank top row
        next(rows, None)  # header: Source, Domain, Subdomain, Control ID, Control Statement, Requirements/Evidence

        domain_cache = {}   # (fw_code, domain_name) -> Domain
        created = updated = skipped = 0
        per_fw = {}

        for r in rows:
            if not r:
                continue
            src = _clean(r[0])
            domain_raw = _clean(r[1])
            subdomain = _clean(r[2])
            cid = _clean(r[3])
            statement = _clean(r[4])
            req = _clean(r[5])
            if not cid:
                continue
            fw = FRAMEWORKS.get(src)
            if not fw:
                skipped += 1
                continue
            fw_code = fw['code']
            framework = fw_objs[fw_code]

            # Domain = subdomain text (reliable) else top-level domain else General
            domain_name = subdomain or domain_raw or 'General'
            domain_name = domain_name[:200]
            dkey = (fw_code, domain_name)
            domain = domain_cache.get(dkey)
            if domain is None:
                code_token = (domain_name.split()[0] if domain_name.split() else 'GEN')[:50]
                domain, _ = Domain.objects.get_or_create(
                    framework=framework, name=domain_name,
                    defaults={'code': code_token, 'order': len(domain_cache)},
                )
                domain_cache[dkey] = domain

            ev = evidence.get((fw_code, cid), {})
            evidence_desc = ev.get('evidence_desc', '')
            guidance = evidence_desc or req
            ev_type = _evidence_type(evidence_desc) if evidence_desc else 'policy'
            priority = _priority(ev.get('theme', ''))

            obj, was_created = Control.objects.update_or_create(
                framework=framework, control_id=cid,
                defaults={
                    'domain': domain,
                    'title': statement[:500] or cid,
                    'description': statement,
                    'priority': priority,
                    'evidence_type': ev_type,
                    'evidence_guidance': guidance,
                    'evidence_guidance_ar': ev.get('guidance_ar', ''),
                    'is_mandatory': True,
                },
            )
            created += int(was_created)
            updated += int(not was_created)
            per_fw[fw_code] = per_fw.get(fw_code, 0) + 1

        wb.close()

        # 4) Refresh framework totals
        for code, obj in fw_objs.items():
            obj.total_controls = Control.objects.filter(framework=obj).count()
            obj.save(update_fields=['total_controls'])

        total = Control.objects.filter(framework__code__in=[m['code'] for m in FRAMEWORKS.values()]).count()
        self.stdout.write(self.style.SUCCESS(
            f'Seeded controls — created={created}, updated={updated}, skipped={skipped}'
        ))
        for code, n in sorted(per_fw.items()):
            self.stdout.write(f'  {code}: {n}')
        self.stdout.write(self.style.SUCCESS(f'TOTAL controls now: {total}'))
